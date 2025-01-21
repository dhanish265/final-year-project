from numpy import random
from scipy.stats import beta
import xlsxwriter
from openpyxl import load_workbook
import os.path
import bisect
from utils import MIN_IDEAL_DIST, MAX_IDEAL_DIST

LIMIT = 21 * 24 * 60
dwell_times = [None] * 3 #normal, idle, busy
boundaries = [None] * 3
dwell_time_raw_data = None
exp_rate = 0.45
LOWER_LIMIT, UPPER_LIMIT = 24 * 60, 24 * 60 * 2
# print(times)

def generateData(busy_rate = 1, name = 'Synthetic Anchorage (normal)'):
    # workbook = xlsxwriter.Workbook(name  + '.xlsx')
    data = []
    for _ in range(1):
        # worksheet = workbook.add_worksheet(str(i + 1))
        data_sheet = []
        currTime = random.exponential(exp_rate) * 60
        while currTime < LIMIT:
            data_sheet.append((round(currTime), round(currTime) + round(busy_rate * random.lognormal(2.4, 1.3) * 60), beta.rvs(2.4, 2.4, 30, 270).item()))
            currTime += random.exponential(exp_rate) * 60
        data.append(data_sheet)
    write_to_xlsx(data, name + '.xlsx', [str(i) for i in range(1, 101)])

    
def read_data(name):
    print("data/" + name + ".xlsx")
    if not os.path.isfile("data/" + name + ".xlsx"):
        print("Incorrect file name! Please check!")
        return None
    data = {}
    workbook = load_workbook("data/" + name + ".xlsx")
    sheetNames = workbook.sheetnames
    for sheetName in sheetNames:
        values = []
        worksheet = workbook[sheetName]
        for row in worksheet.values:
            values.append(row)
        data[sheetName] = values
    return data

# data = read_data('Ahirkapi Anchorage.xlsx')
# if data is not None:
#     print(data['100'])
# generateData(busy_rate= 0.5, name = 'Synthetic Anchorage (idle)')

def dwell_time_dist_analysis(rewrite = False, busy_rate = 1, sheet_name = 'normal'):
    global dwell_time_raw_data
    value = []
    if not rewrite:
        if dwell_time_raw_data is None:
            dwell_time_raw_data = read_data('dwell_time_analysis')
        for row in dwell_time_raw_data[sheet_name]:
            value.append(row[-1])
        return value
    
    data = [[] for i in range(730)]
    for _ in range(100):
        data2 = []
        for j in range(730):
            data2.append(round(busy_rate * random.lognormal(2.4, 1.3) * 60))
        data2.sort()
        for j in range(730):
            # print(j)
            data[j].append(data2[j])
    for j in range(730):
        average = sum(data[j])/len(data[j])
        data[j].append(average)
        value.append(average)
    write_to_xlsx([data], 'dwell_time_analysis', [sheet_name]) 
    return value

def write_to_xlsx(data, book_name: str, sheet_names: list[str]):
    workbook = xlsxwriter.Workbook(book_name)
    for i, data_sheet in enumerate(data):
        worksheet = workbook.add_worksheet(sheet_names[i])
        row = 0
        while row < len(data_sheet):
            col = 0
            while col < len(data_sheet[row]):
                worksheet.write(row, col, data_sheet[row][col])
                col += 1
            row += 1
    workbook.close()


def instantiate_times():
    print('instantiation done')
    dwell_times[0] = dwell_time_dist_analysis(rewrite = False, busy_rate=1, sheet_name='normal')
    dwell_times[1] = dwell_time_dist_analysis(rewrite = False, busy_rate=0.5, sheet_name='idle')
    dwell_times[2] = dwell_time_dist_analysis(rewrite = False, busy_rate=2.2, sheet_name='busy')
    boundaries[0] = [(dwell_times[0][i] + dwell_times[0][i+1])/2 for i in range(len(dwell_times[0]) - 1)]
    boundaries[1] = [(dwell_times[1][i] + dwell_times[1][i+1])/2 for i in range(len(dwell_times[1]) - 1)]
    boundaries[2] = [(dwell_times[2][i] + dwell_times[2][i+1])/2 for i in range(len(dwell_times[2]) - 1)]
    # the times are obtained as [a, b, c, d, ... y, z]
    # using this, if the dwell time of a vessel is between 0 and (a+b)/2, it's ideal anchoring distance from entry side is 445
    # likewise, if it is > (y+z)/2, it's ideal anchoring distance is 2500 - 445 = 2055
    # everything in between is evenly distributed between 445 and 2055 (i.e., if dwell time is in between (a+b)/2 and (b+c)/2, 445 + (2025-445)/(26-1))

def obtain_ideal_distance(dwell_time, busy_status = 'normal'):
    sheet_list = {'normal': 0, 'idle': 1, 'busy': 2}
    index = sheet_list[busy_status]
    return MIN_IDEAL_DIST + bisect.bisect(boundaries[index], dwell_time) * (MAX_IDEAL_DIST - MIN_IDEAL_DIST)/(len(boundaries[index]))

# instantiate_times()
# obtain_ideal_distance(19.52, 'busy')
# generateData(name='test_0.45_inverse')



